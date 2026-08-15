import datetime as dt
import random

from openpyxl import Workbook

from src.ingestion.excel_connector import ExcelConnector
from src.simulator.amazon_simulator import generate_amazon_batch
from src.simulator.universe import Universe


def test_excel_connector_reads_all_sheets(tmp_path):
    universe = Universe(seed=1)
    path = generate_amazon_batch(universe, dt.date(2024, 3, 1), 15, tmp_path,
                                  rng=random.Random(1))
    conn = ExcelConnector(path)
    conn.connect()
    sheets = conn.extract()
    assert set(sheets.keys()) == {"Orders", "Returns", "Fee Breakdown", "Adjustments"}
    assert len(sheets["Orders"]) == 15


def test_excel_connector_single_sheet(tmp_path):
    universe = Universe(seed=1)
    path = generate_amazon_batch(universe, dt.date(2024, 3, 1), 10, tmp_path,
                                  rng=random.Random(2))
    conn = ExcelConnector(path, sheet_name="Orders")
    conn.connect()
    df = conn.extract()
    assert len(df) == 10


def test_excel_connector_fuzzy_sheet_name_match(tmp_path):
    universe = Universe(seed=1)
    path = generate_amazon_batch(universe, dt.date(2024, 3, 1), 5, tmp_path,
                                  rng=random.Random(3))
    conn = ExcelConnector(path, sheet_name="order")  # lowercase substring
    conn.connect()
    df = conn.extract()
    assert len(df) == 5


def test_excel_connector_unmerges_cells(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Header1", "Header2"])
    ws.append(["merged_value", None])
    ws.merge_cells("A2:A3")
    ws.append([None, "row3_b"])
    path = tmp_path / "merged.xlsx"
    wb.save(path)

    conn = ExcelConnector(path, sheet_name="Sheet")
    conn.connect()
    df = conn.extract()
    assert df.iloc[0]["Header1"] == "merged_value"
    assert df.iloc[1]["Header1"] == "merged_value"  # filled from merged range


def test_excel_connector_converts_formula_errors_to_na(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append([1, "#REF!"])
    ws.append([2, "#DIV/0!"])
    path = tmp_path / "errors.xlsx"
    wb.save(path)

    conn = ExcelConnector(path, sheet_name="Sheet")
    conn.connect()
    df = conn.extract()
    assert df["b"].isna().sum() == 2
