"""Excel connector: multi-sheet reading, merged-cell unmerging, and
Excel formula-error detection (#REF!, #DIV/0!, #N/A -> NaN).
"""

from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from src.ingestion.base_connector import BaseConnector

EXCEL_ERROR_VALUES = {"#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NAME?", "#NULL!", "#NUM!"}


def _unmerge_sheet_values(ws) -> None:
    """Fills every cell in a merged range with the top-left cell's value,
    the way pandas would expect a normal (non-merged) sheet to look."""
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value


def _sheet_to_dataframe(ws) -> pd.DataFrame:
    _unmerge_sheet_values(ws)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    header, *data_rows = rows
    header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]
    df = pd.DataFrame(data_rows, columns=header)
    df = df.replace(list(EXCEL_ERROR_VALUES), pd.NA)
    return df


class ExcelConnector(BaseConnector):
    def __init__(self, source_path, sheet_name: str | None = None):
        super().__init__(source_path)
        self.sheet_name = sheet_name
        self.formula_error_count = 0

    def connect(self) -> None:
        if not self.source_path.exists():
            raise FileNotFoundError(self.source_path)
        self._connected = True

    def extract(self) -> pd.DataFrame | dict[str, pd.DataFrame]:
        wb = load_workbook(self.source_path, data_only=True)
        try:
            if self.sheet_name:
                return _sheet_to_dataframe(wb[self._resolve_sheet_name(wb, self.sheet_name)])
            return {name: _sheet_to_dataframe(wb[name]) for name in wb.sheetnames}
        finally:
            wb.close()

    @staticmethod
    def _resolve_sheet_name(wb, requested: str) -> str:
        """Sheet names sometimes vary between files (e.g. 'Orders' vs
        'Order Data'); fall back to a case-insensitive substring match."""
        if requested in wb.sheetnames:
            return requested
        for name in wb.sheetnames:
            if requested.lower() in name.lower():
                return name
        raise KeyError(f"No sheet matching '{requested}' in {wb.sheetnames}")
